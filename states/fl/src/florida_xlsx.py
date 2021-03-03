from openpyxl import load_workbook
import os
import re
from dataclasses import (
    dataclass,
    asdict as dc2dict,
)
from typing import Any
from datetime import datetime
import json

SOURCE_URL = \
    "https://dos.myflorida.com/media/703568/" + \
    "excel-early-voting-locations-for-2020-general.zip"

XLSX_PATH = os.path.join(
    "..", "raw_data", "Excel Early Voting Locations for 2020 General"
)

OUTPUT_PATH = os.path.join(
    "..", "raw_data", "all-florida-stage1.json"
)

YEAR = 2020

loc_rx = re.compile(r'^Location \d+$')
zip_rx = re.compile(r'\d{5}')

@dataclass
class Sched:
    startDate: Any = None
    endDate: Any = None
    startTime: Any = None
    endTime: Any = None

@dataclass
class Location:
    raw_address: str
    county: str
    municipality: str
    name: str
    source: str
    state: str
    stateAbbreviation: str
    type: str
    zip: Any = None
    timezone: Any = None
    city: Any = None
    notes: Any = None
    phone: Any = None
    schedule: Any = None

def load_excel_files():
    return {
        xl_path: load_workbook(
            os.path.join(XLSX_PATH, xl_path), data_only=True
        ) 
        for xl_path in filter(
            lambda x: x[-4:] == 'xlsx', 
            list(os.listdir(XLSX_PATH))
        )
    }

def parse_counties(excel_files):
    counties = dict()
    for fn, d in excel_files.items():
        sheetname = list(filter(lambda s: s.lower().startswith('general'), d.sheetnames))[0]
        sheet = d[sheetname]
        county = " ".join(sheet.cell(3,1).value.split()[:-1])
        locations = parse_county_locations(county, sheet)
        counties[county] = {
            "sheet": d, 
            "filename": fn, 
            "sheet": sheet, 
            "locations": locations
        }
    return counties

def parse_county_locations(county_name, sheet):
    loc_start_rows = []
    for i, r in enumerate(sheet.rows):
        if loc_rx.match(str(r[0].value)):
            loc_start_rows.append(i + 1)
    locations = [
        parse_location(county_name, sheet, i)
        for i in loc_start_rows
    ]
    return locations

def parse_location(county, sheet, start_row_i):
    data_i = start_row_i + 2
    name = sheet.cell(data_i, 1).value
    loc_type = sheet.cell(data_i, 2).value
    address = sheet.cell(data_i, 3).value
    timezone = sheet.cell(data_i, 7).value
    try:
        zip = zip_rx.findall(address)[-1]
    except:
        zip = None
    sched = parse_sched(sheet, start_row_i + 3)
    l = Location(
        raw_address=address,
        county=county,
        municipality=county,
        name=name,
        source=SOURCE_URL,
        state="Florida",
        stateAbbreviation="FL",
        type="earlyVoting",
        zip=zip,
        timezone=timezone,
        schedule=sched,
    )
    return l


def parse_sched(sheet, start_row_i):
    i = start_row_i
    schedules = []
    while sheet.cell(i, 1).value is None or sheet.cell(i, 1).value == "Days of Operation":
        i += 1
    if not 'oct' in sheet.cell(i, 1).value.lower():
        print("something funky with the date, skipping")
    while sheet.cell(i, 1).value is not None:
        try:
            is_avail = str(sheet.cell(i, 3).value).lower()[0] == "y"
            if is_avail:
                dt = datetime.strptime(
                    sheet.cell(i, 1).value.split("(")[0].strip(),
                    "%A, %B %d"
                ).strftime("%-m/%-d/{}".format(YEAR))
                start = str(sheet.cell(i, 4).value)
                end = str(sheet.cell(i, 6).value)
                schedules.append(
                    Sched(dt, dt, start, end)
                )
        except ValueError:
            pass
        i += 1
    return schedules

if __name__ == "__main__":
    excel_files = load_excel_files()
    county_data = parse_counties(excel_files)
    all_locations = []
    for c in county_data.values():
        all_locations.extend(c["locations"])
    with open(OUTPUT_PATH, "w") as f:
        json.dump([dc2dict(l) for l in all_locations], f, indent=2)